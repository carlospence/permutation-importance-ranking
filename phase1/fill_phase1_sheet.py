"""
Fills CS6735_PROJECT_-_PHASE_1.xlsx with Phase 1 results.
- Fold values filled into rows 3-192 (template untouched structurally)
- Row 193: single mean ± std row, one value per metric column
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

TEMPLATE   = "CS6735_PROJECT_-_PHASE_1.xlsx"
OUTPUT     = "CS6735_PROJECT_-_PHASE_1_FILLED.xlsx"
RESULTS    = "results/phase1"
N_DATASETS = 16
N_FOLDS    = 10

COLS = {
    "svm": {"acc": 2,  "f1": 3,  "par": 12},
    "knn": {"acc": 4,  "f1": 5,  "par": 13},
    "dt":  {"acc": 6,  "f1": 7,  "par": 14},
    "rf":  {"acc": 8,  "f1": 9,  "par": 15},
    "mlp": {"acc": 10, "f1": 11, "par": 16},
}

def fold_start_row(dataset_num: int) -> int:
    return 3 + (dataset_num - 1) * 12

def load_folds(clf: str, dataset_num: int) -> pd.DataFrame | None:
    path = os.path.join(RESULTS, clf, f"dataset_{dataset_num}_{clf}_folds.csv")
    if not os.path.exists(path):
        print(f"  [!] Missing: {path}")
        return None
    return pd.read_csv(path).sort_values("Fold").reset_index(drop=True)


def main():
    if not os.path.exists(TEMPLATE):
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")

    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # Collect all values per column for final mean±std row
    all_values = {clf: {"acc": [], "f1": []} for clf in COLS}

    # ── Fill fold values ───────────────────────────────────────────────────────
    for dataset_num in range(1, N_DATASETS + 1):
        start = fold_start_row(dataset_num)
        print(f"Dataset {dataset_num} — rows {start} to {start + N_FOLDS - 1}")

        for clf, cols in COLS.items():
            folds_df = load_folds(clf, dataset_num)
            if folds_df is None:
                continue

            for fold_idx in range(N_FOLDS):
                row  = start + fold_idx
                data = folds_df[folds_df["Fold"] == fold_idx + 1]
                if data.empty:
                    continue

                acc = round(float(data["Accuracy"].values[0]), 4)
                f1  = round(float(data["F1"].values[0]), 4)
                par = str(data["Parameters"].values[0])

                ws.cell(row=row, column=cols["acc"]).value = acc
                ws.cell(row=row, column=cols["f1"]).value  = f1
                ws.cell(row=row, column=cols["par"]).value = par

                all_values[clf]["acc"].append(acc)
                all_values[clf]["f1"].append(f1)

            print(f"  ✓ {clf.upper()}")

    # ── Single mean ± std row at row 193 ──────────────────────────────────────
    summary_row  = 193
    green_fill   = PatternFill("solid", start_color="D9EAD3")
    bold_font    = Font(bold=True)
    center_align = Alignment(horizontal="center")

    ws.cell(row=summary_row, column=1).value = "Mean ± Std"
    ws.cell(row=summary_row, column=1).font  = bold_font
    ws.cell(row=summary_row, column=1).fill  = green_fill

    for clf, cols in COLS.items():
        acc_vals = all_values[clf]["acc"]
        f1_vals  = all_values[clf]["f1"]

        if acc_vals:
            acc_mean = np.mean(acc_vals)
            acc_std  = np.std(acc_vals, ddof=1)
            f1_mean  = np.mean(f1_vals)
            f1_std   = np.std(f1_vals, ddof=1)

            for col, val in [
                (cols["acc"], f"{acc_mean:.4f} ± {acc_std:.4f}"),
                (cols["f1"],  f"{f1_mean:.4f} ± {f1_std:.4f}"),
            ]:
                cell            = ws.cell(row=summary_row, column=col)
                cell.value      = val
                cell.font       = bold_font
                cell.fill       = green_fill
                cell.alignment  = center_align

    # Widen columns so mean±std text fits
    for clf, cols in COLS.items():
        for col in [cols["acc"], cols["f1"]]:
            letter = ws.cell(row=summary_row, column=col).column_letter
            ws.column_dimensions[letter].width = 22
    ws.column_dimensions["A"].width = 14

    wb.save(OUTPUT)
    print(f"\nSaved → {OUTPUT}")
    print(f"Mean ± Std row written at row {summary_row}")


if __name__ == "__main__":
    main()