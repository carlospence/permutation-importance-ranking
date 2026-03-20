"""
Fills both sheets in the project workbook:
  - "Before FS-DR" sheet → Phase 1 results
  - "After FS-DR"  sheet → Phase 2 results

Usage:
  python fill_sheets.py

Place this script in your project root alongside the workbook.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment

# ── Config ─────────────────────────────────────────────────────────────────────

WORKBOOK     = "CS6735_PROJECT.xlsx"        # rename your workbook to this
PHASE1_SHEET = "Before FS-DR"
PHASE2_SHEET = "After FS-DR"
PHASE1_DIR   = "results/phase1"
PHASE2_DIR   = "results/phase2"
N_DATASETS   = 16
N_FOLDS      = 10
TARGET_COL   = "Label"

FS_PARAMS = "RF permutation importance | n_repeats=10 | threshold=mean importance"

# Phase 1 column indices (1-based)
P1_COLS = {
    "svm": {"acc": 2,  "f1": 3,  "par": 12},
    "knn": {"acc": 4,  "f1": 5,  "par": 13},
    "dt":  {"acc": 6,  "f1": 7,  "par": 14},
    "rf":  {"acc": 8,  "f1": 9,  "par": 15},
    "mlp": {"acc": 10, "f1": 11, "par": 16},
}

# Phase 2 column indices (1-based)
P2_COLS = {
    "svm": {"acc": 2,  "f1": 3,  "par": 14},
    "knn": {"acc": 4,  "f1": 5,  "par": 15},
    "dt":  {"acc": 6,  "f1": 7,  "par": 16},
    "rf":  {"acc": 8,  "f1": 9,  "par": 17},
    "mlp": {"acc": 10, "f1": 11, "par": 18},
}
P2_COL_N_FEATURES = 12
P2_COL_FEAT_NAMES = 13


# ── Helpers ────────────────────────────────────────────────────────────────────

def fold_start_row(dataset_num: int) -> int:
    """Fold 1 row for dataset N — pattern: 3 + (N-1) * 12."""
    return 3 + (dataset_num - 1) * 12


def write_cell(ws, row, col, value):
    """Write to cell only if it's not a merged cell."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def write_summary_row(ws, row, cols_dict, all_values):
    """Write mean ± std row at the given row number."""
    green_fill   = PatternFill("solid", start_color="D9EAD3")
    bold_font    = Font(bold=True)
    center_align = Alignment(horizontal="center")

    cell = ws.cell(row=row, column=1)
    if not isinstance(cell, MergedCell):
        cell.value = "Mean ± Std"
        cell.font  = bold_font
        cell.fill  = green_fill

    for clf, cols in cols_dict.items():
        acc_vals = all_values[clf]["acc"]
        f1_vals  = all_values[clf]["f1"]
        if not acc_vals:
            continue

        acc_mean = np.mean(acc_vals)
        acc_std  = np.std(acc_vals, ddof=1)
        f1_mean  = np.mean(f1_vals)
        f1_std   = np.std(f1_vals, ddof=1)

        for col, val in [
            (cols["acc"], f"{acc_mean:.4f} ± {acc_std:.4f}"),
            (cols["f1"],  f"{f1_mean:.4f} ± {f1_std:.4f}"),
        ]:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value     = val
            cell.font      = bold_font
            cell.fill      = green_fill
            cell.alignment = center_align


def auto_width(ws, cols, width=22):
    from openpyxl.utils import get_column_letter
    for col in cols:
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Phase 1 filler ─────────────────────────────────────────────────────────────

def fill_phase1(ws):
    print(f"\n{'='*50}")
    print(f"Filling: {PHASE1_SHEET}")
    print(f"{'='*50}")

    all_values = {clf: {"acc": [], "f1": []} for clf in P1_COLS}

    for dataset_num in range(1, N_DATASETS + 1):
        start = fold_start_row(dataset_num)
        print(f"\nDataset {dataset_num} — rows {start} to {start + N_FOLDS - 1}")

        for clf, cols in P1_COLS.items():
            path = os.path.join(PHASE1_DIR, clf, f"dataset_{dataset_num}_{clf}_folds.csv")
            if not os.path.exists(path):
                print(f"  [!] Missing: {path}")
                continue

            folds_df = pd.read_csv(path).sort_values("Fold").reset_index(drop=True)

            for fold_idx in range(N_FOLDS):
                row  = start + fold_idx
                data = folds_df[folds_df["Fold"] == fold_idx + 1]
                if data.empty:
                    continue

                acc = round(float(data["Accuracy"].values[0]), 4)
                f1  = round(float(data["F1"].values[0]), 4)
                par = str(data["Parameters"].values[0])

                write_cell(ws, row, cols["acc"], acc)
                write_cell(ws, row, cols["f1"],  f1)
                write_cell(ws, row, cols["par"], par)

                all_values[clf]["acc"].append(acc)
                all_values[clf]["f1"].append(f1)

            print(f"  ✓ {clf.upper()}")

    write_summary_row(ws, 193, P1_COLS, all_values)
    auto_width(ws, range(2, 17))
    ws.column_dimensions["A"].width = 14
    print(f"\nPhase 1 done — Mean ± Std at row 193")


# ── Phase 2 filler ─────────────────────────────────────────────────────────────

def fill_phase2(ws):
    print(f"\n{'='*50}")
    print(f"Filling: {PHASE2_SHEET}")
    print(f"{'='*50}")

    all_values = {clf: {"acc": [], "f1": []} for clf in P2_COLS}

    for dataset_num in range(1, N_DATASETS + 1):
        start = fold_start_row(dataset_num)
        path  = os.path.join(PHASE2_DIR, f"dataset_{dataset_num}_all_p2_folds.csv")

        if not os.path.exists(path):
            print(f"  [!] Missing: {path} — skipping dataset {dataset_num}")
            continue

        folds_df = pd.read_csv(path)
        print(f"\nDataset {dataset_num} — rows {start} to {start + N_FOLDS - 1}")

        # Shared columns from SVM folds (features same across classifiers)
        svm_folds = folds_df[folds_df["Classifier"] == "SVM"].sort_values("Fold").reset_index(drop=True)

        for fold_idx in range(N_FOLDS):
            row = start + fold_idx

            if fold_idx < len(svm_folds):
                n_feat     = svm_folds.iloc[fold_idx].get("N_Features_After", "")
                feat_names = svm_folds.iloc[fold_idx].get("Selected_Features", "")
                write_cell(ws, row, P2_COL_N_FEATURES, int(n_feat) if pd.notna(n_feat) else "")
                write_cell(ws, row, P2_COL_FEAT_NAMES, str(feat_names) if pd.notna(feat_names) else "")

            for clf, cols in P2_COLS.items():
                clf_folds = folds_df[folds_df["Classifier"] == clf.upper()].sort_values("Fold").reset_index(drop=True)
                if fold_idx >= len(clf_folds):
                    continue

                fold_data = clf_folds.iloc[fold_idx]
                acc = round(float(fold_data["Accuracy"]), 4)
                f1  = round(float(fold_data["F1"]), 4)

                write_cell(ws, row, cols["acc"], acc)
                write_cell(ws, row, cols["f1"],  f1)
                write_cell(ws, row, cols["par"], FS_PARAMS)

                all_values[clf]["acc"].append(acc)
                all_values[clf]["f1"].append(f1)

        print(f"  ✓ All classifiers written")

    write_summary_row(ws, 193, P2_COLS, all_values)
    auto_width(ws, range(2, 19))
    ws.column_dimensions["A"].width = 14
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(P2_COL_FEAT_NAMES)].width = 40
    print(f"\nPhase 2 done — Mean ± Std at row 193")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(WORKBOOK):
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK}\n"
            f"Rename your project workbook to '{WORKBOOK}' and place it in the project root."
        )

    wb = load_workbook(WORKBOOK)

    if PHASE1_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{PHASE1_SHEET}' not found. Available: {wb.sheetnames}")
    if PHASE2_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{PHASE2_SHEET}' not found. Available: {wb.sheetnames}")

    fill_phase1(wb[PHASE1_SHEET])
    fill_phase2(wb[PHASE2_SHEET])

    wb.save(WORKBOOK)
    print(f"\nSaved → {WORKBOOK}")


if __name__ == "__main__":
    main()